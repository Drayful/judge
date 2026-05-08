import json
import re
from datetime import date, datetime
from pathlib import Path


XLSX_PATH = Path(r"C:\Users\Admin\Downloads\grad-2014.xlsx")
OUTPUT_PATH = Path(r"C:\Users\Admin\Desktop\grad-2014.json")


def is_all_zeros_format(fmt: str) -> int | None:
    """
    Detect simple leading-zero numeric formats like '000000' (optionally with spaces).
    Returns width if matches.
    """
    if not fmt:
        return None
    fmt = fmt.strip()
    if re.fullmatch(r"0+", fmt):
        return len(fmt)
    return None


def normalize_numeric_string(s: str) -> str:
    """
    Turn '21.0' -> '21', '0.0' -> '0'. Keep other strings intact.
    """
    s2 = s.strip()
    if re.fullmatch(r"-?\d+\.0", s2):
        return s2[:-2]
    return s


def to_ymd(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def format_cell_value(cell) -> object:
    """
    Preserve leading zeros if the cell has a pure-zero number format like '0000000'.
    Also normalize dates to YYYY-MM-DD and convert integer-ish floats to ints.
    """
    v = cell.value
    if v is None:
        return None

    d = to_ymd(v)
    if d is not None:
        return d

    # keep booleans as is
    if isinstance(v, bool):
        return v

    # strings: keep, but normalize trailing .0 if present
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
        return normalize_numeric_string(v)

    # numeric: preserve with leading zeros if format requires it
    if isinstance(v, (int, float)):
        # if float is integer-ish, treat as int for formatting purposes
        is_intish = isinstance(v, int) or (isinstance(v, float) and float(v).is_integer())
        if is_intish:
            iv = int(v)
            width = is_all_zeros_format(getattr(cell, "number_format", "") or "")
            if width:
                return str(iv).zfill(width)
            return iv
        # non-integer float: keep as-is
        return v

    return v


def drop_unnamed(headers: list[str], row_dict: dict) -> dict:
    return {
        k: v
        for k, v in row_dict.items()
        if k is not None
        and str(k).strip() != ""
        and not str(k).strip().startswith("Unnamed")
        and str(k) in headers
    }


def main() -> None:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise SystemExit(f"openpyxl is required: {e}")

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.worksheets[0]

    rows = list(ws.iter_rows())
    if not rows:
        OUTPUT_PATH.write_text("[]", encoding="utf-8")
        print(f"Wrote empty array -> {OUTPUT_PATH}")
        return

    headers = []
    for c in rows[0]:
        h = c.value
        headers.append(str(h).strip() if h is not None else "")

    out = []
    for r in rows[1:]:
        obj = {}
        empty = True
        for idx, cell in enumerate(r):
            if idx >= len(headers):
                break
            key = headers[idx]
            if not key or key.startswith("Unnamed"):
                continue
            val = format_cell_value(cell)
            if val is not None:
                empty = False
            obj[key] = val
        if empty:
            continue

        # business rules carried over from earlier steps
        if obj.get("centerProfChecked") is True:
            obj["professionId"] = 0

        out.append(obj)

    OUTPUT_PATH.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(out)} rows -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

